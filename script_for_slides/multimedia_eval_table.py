#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import json
import argparse
import re
from typing import Dict, Tuple, Optional, List

import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


DEFAULT_THRESHOLD = 0.75


# ----------------- I/O -----------------

def read_json_or_jsonl(path: str) -> dict:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".json":
        with open(path, "r") as f:
            return json.load(f)
    if ext == ".jsonl":
        with open(path, "r") as f:
            for line in f:
                line = line.strip()
                if line:
                    return json.loads(line)
        raise ValueError("Fichier .jsonl vide")
    with open(path, "r") as f:
        return json.load(f)


def extract_final_dict(payload: dict) -> Dict[str, float]:
    results = payload.get("results", {})
    final_dict = results.get("final_dict", None)
    if not isinstance(final_dict, dict):
        raise KeyError("Impossible de trouver results.final_dict (dict) dans le JSON.")
    return final_dict


# ----------------- Parsing / sorting -----------------

def parse_attack_name(name: str) -> Tuple[str, Optional[float], str]:
    """
    Retourne (attack_type, numeric_severity, raw_tail)
    Ex:
      crop_05 -> ("crop", 0.5, "05")
      crop_01 -> ("crop", 0.1, "01")
      rot_25 -> ("rot", 25.0, "25")
      brightness_1p5 -> ("brightness", 1.5, "1p5")
    """
    if name == "none":
        return ("none", None, "")

    parts = name.split("_")
    attack_type = parts[0]

    if len(parts) >= 2:
        tail = "_".join(parts[1:])
        nums = re.findall(r"\d+p\d+|\d+\.?\d*", tail)
        if nums:
            token = nums[0].replace("p", ".")
            try:
                val = float(token)
                # heuristique: "05"/"01" dans crop/resize => 0.5/0.1
                if re.fullmatch(r"\d{2}", nums[0]) and attack_type in {"crop", "resize"} and val >= 1.0:
                    val = val / 10.0
                return (attack_type, val, tail)
            except ValueError:
                pass
        return (attack_type, None, tail)

    return (attack_type, None, "")


def attack_type_order_key(attack_type: str) -> int:
    # ordre lisible (ajuste si tu veux)
    order = [
        "none",
        "crop",
        "resize",
        "rot",
        "jpeg",
        "brightness",
        "contrast",
        "saturation",
        "sharpness",
        "overlay",
        "comb",
    ]
    if attack_type in order:
        return order.index(attack_type)
    return len(order) + 1


def sort_attacks(names: List[str]) -> List[str]:
    """
    - none en premier
    - groupé par type (ordre custom)
    - à l'intérieur :
        * jpeg : qualité plus haute d'abord (80 -> 50)
        * autres numériques : croissant (0.1 -> 0.5, 25 -> 90, 1.5 -> 2)
        * sinon lexicographique
    """
    def key(n: str):
        at, sev, _ = parse_attack_name(n)
        if n == "none":
            return (0, 0, 0, n)

        type_rank = attack_type_order_key(at)

        if at == "jpeg" and sev is not None:
            return (1, type_rank, -sev, n)

        if sev is not None:
            return (1, type_rank, sev, n)

        return (1, type_rank, float("inf"), n)

    return sorted(names, key=key)


# ----------------- Pretty labels -----------------

def fmt_num(x: float) -> str:
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    return f"{x:.3f}".rstrip("0").rstrip(".")


def pretty_attack_label(raw: str) -> str:
    """
    Ex:
      crop_01 -> "Crop: 0.1"
      rot_25 -> "Rotation: 25°"
      brightness_1p5 -> "Brightness: ×1.5"
      overlay_text -> "Overlay: text"
      none -> "Baseline (none)"
    """
    if raw == "none":
        return "Baseline (none)"

    at, sev, tail = parse_attack_name(raw)
    at_low = at.lower()

    name_map = {
        "crop": "Crop",
        "resize": "Resize",
        "rot": "Rotation",
        "jpeg": "JPEG",
        "brightness": "Brightness",
        "contrast": "Contrast",
        "saturation": "Saturation",
        "sharpness": "Sharpness",
        "overlay": "Overlay",
        "comb": "Combined",
    }
    pretty_type = name_map.get(at_low, at)

    if at_low == "overlay":
        # ex: overlay_text
        if "text" in raw:
            return "Overlay: text"
        return "Overlay"

    if at_low == "comb":
        return "Combined"

    if sev is not None:
        if at_low in {"crop", "resize"}:
            return f"{pretty_type}: {fmt_num(sev)}"
        if at_low == "rot":
            return f"{pretty_type}: {fmt_num(sev)}°"
        if at_low == "jpeg":
            return f"{pretty_type}: {fmt_num(sev)}"
        if at_low in {"brightness", "contrast", "saturation", "sharpness"}:
            return f"{pretty_type}: ×{fmt_num(sev)}"
        return f"{pretty_type}: {fmt_num(sev)}"

    # fallback
    return raw.replace("_", " ")


# ----------------- LaTeX escaping -----------------

def latex_escape(s: str) -> str:
    replacements = [
        ("\\", r"\textbackslash{}"),
        ("&", r"\&"),
        ("%", r"\%"),
        ("$", r"\$"),
        ("#", r"\#"),
        ("_", r"\_"),
        ("{", r"\{"),
        ("}", r"\}"),
        ("~", r"\textasciitilde{}"),
        ("^", r"\textasciicircum{}"),
    ]
    for orig, repl in replacements:
        s = s.replace(orig, repl)
    return s


# ----------------- Writers -----------------

def save_latex(attacks: List[str], values: Dict[str, float], out_path: str, caption: str, threshold: float):
    with open(out_path, "w") as f:
        f.write("%% Auto-generated table\n")
        f.write("%% Requires in your preamble:\n")
        f.write("%%   \\usepackage[table]{xcolor}\n")
        f.write("\\begin{table}[ht]\n\\centering\n")
        f.write("\\begin{tabular}{lc}\n")
        f.write("\\hline\n")
        f.write("Attack & Bit acc. ($\\uparrow$) \\\\\n")
        f.write("\\hline\n")

        for a in attacks:
            v = values.get(a, None)
            a_tex = latex_escape(pretty_attack_label(a))
            if isinstance(v, (float, int)):
                v_str = f"{float(v):.4f}"
                if float(v) >= threshold:
                    v_str = "\\cellcolor{green!20} " + v_str
            else:
                v_str = latex_escape(str(v)) if v is not None else ""
            f.write(f"{a_tex} & {v_str} \\\\\n")

        f.write("\\hline\n")
        f.write("\\caption{%s}\n" % latex_escape(caption))
        f.write("\\end{tabular}\n\\end{table}\n")


def save_png(attacks: List[str], values: Dict[str, float], out_path: str, title: str, threshold: float):
    rows = []
    for a in attacks:
        v = values.get(a, "")
        rows.append([pretty_attack_label(a), f"{float(v):.4f}" if isinstance(v, (float, int)) else str(v)])

    plt.rcParams["font.family"] = "serif"
    plt.rcParams["font.size"] = 10

    fig_w = 7.5
    fig_h = max(2.8, 0.35 * (len(rows) + 1))

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")

    table = ax.table(
        cellText=rows,
        colLabels=["Attack", "Bit acc. (↑)"],
        loc="center",
        cellLoc="center",
        colLoc="center",
    )

    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1.0, 1.2)

    for c in range(2):
        table[(0, c)].set_text_props(weight="bold")

    for i, a in enumerate(attacks, start=1):
        v = values.get(a, None)
        if isinstance(v, (float, int)) and float(v) >= threshold:
            table[(i, 1)].set_facecolor((0.85, 0.95, 0.85))

    ax.set_title(title, pad=10)
    plt.tight_layout()
    plt.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close()


def save_xlsx(attacks: List[str], values: Dict[str, float], out_path: str, threshold: float, sheet_name: str = "multimedia"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    align_center = Alignment(horizontal="center")

    ws.append(["Attack", "Bit accuracy"])
    ws["A1"].font = header_font
    ws["B1"].font = header_font
    ws.freeze_panes = "A2"

    for a in attacks:
        v = values.get(a, None)
        ws.append([pretty_attack_label(a), float(v) if isinstance(v, (float, int)) else ("" if v is None else str(v))])

    for r in range(2, 2 + len(attacks)):
        ws[f"A{r}"].alignment = align_center
        ws[f"B{r}"].alignment = align_center

        cell = ws[f"B{r}"]
        if isinstance(cell.value, (float, int)):
            cell.number_format = "0.0000"
            if float(cell.value) >= threshold:
                cell.fill = green_fill

    for col in range(1, 3):
        col_letter = get_column_letter(col)
        max_len = 0
        for cell in ws[col_letter]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(55, max(12, max_len + 2))

    wb.save(out_path)


# ----------------- Main -----------------

def main():
    parser = argparse.ArgumentParser(
        description="Génère LaTeX/PNG/XLSX à partir d'un JSON contenant results.final_dict (attaque -> bit accuracy)."
    )
    parser.add_argument("json_file", help="Chemin vers le fichier .json (ou .jsonl)")
    parser.add_argument("--out_dir", default=None, help="Dossier de sortie (défaut: dossier du json)")
    parser.add_argument("--threshold", type=float, default=DEFAULT_THRESHOLD, help="Seuil pour coloration verte (défaut: 0.75)")
    args = parser.parse_args()

    payload = read_json_or_jsonl(args.json_file)
    final_dict = extract_final_dict(payload)

    attacks = sort_attacks(list(final_dict.keys()))

    out_dir = args.out_dir or os.path.dirname(os.path.abspath(args.json_file))
    os.makedirs(out_dir, exist_ok=True)

    base = os.path.splitext(os.path.basename(args.json_file))[0]
    tex_path = os.path.join(out_dir, f"{base}_multimedia_table.tex")
    png_path = os.path.join(out_dir, f"{base}_multimedia_table.png")
    xlsx_path = os.path.join(out_dir, f"{base}_multimedia_table.xlsx")

    title = payload.get("metric", "Extraction bit accuracy (multimedia attacks)")
    caption = "Extraction bit accuracy under multimedia attacks (green: >= %.2f)." % float(args.threshold)

    save_latex(attacks, final_dict, tex_path, caption=caption, threshold=float(args.threshold))
    save_png(attacks, final_dict, png_path, title=title, threshold=float(args.threshold))
    save_xlsx(attacks, final_dict, xlsx_path, threshold=float(args.threshold), sheet_name="multimedia")

    print("OK")
    print("LaTeX :", tex_path)
    print("PNG   :", png_path)
    print("XLSX  :", xlsx_path)


if __name__ == "__main__":
    main()
